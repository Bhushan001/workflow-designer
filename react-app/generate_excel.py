#!/usr/bin/env python3
"""
Script to convert estimates.md to an organized Excel file
Improved version with better parsing
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_estimates_md(file_path):
    """Parse the estimates.md file and extract structured data"""
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    phases = []
    current_phase = None
    current_task = None
    in_deliverables = False
    in_breakdown = False
    
    lines = content.split('\n')
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        
        # Phase headers - now supports days
        phase_match = re.match(r'^### Phase (\d+): (.+?) \((\d+)-(\d+)\s*(?:hours|days)\)', line)
        if phase_match:
            if current_phase and current_task:
                current_phase['tasks'].append(current_task)
            if current_phase:
                phases.append(current_phase)
            # Parse as days directly
            min_days = int(phase_match.group(3))
            max_days = int(phase_match.group(4))
            current_phase = {
                'phase_num': int(phase_match.group(1)),
                'phase_name': phase_match.group(2),
                'min_hours': min_days,  # Stored as days now
                'max_hours': max_days,   # Stored as days now
                'tasks': []
            }
            current_task = None
            in_deliverables = False
            in_breakdown = False
            continue
        
        # Task headers
        task_match = re.match(r'^#### Task (\d+\.\d+): (.+?)$', line)
        if task_match and current_phase:
            if current_task:
                current_phase['tasks'].append(current_task)
            current_task = {
                'task_id': task_match.group(1),
                'task_name': task_match.group(2),
                'estimated_time': '',
                'priority': '',
                'breakdown': [],
                'deliverables': []
            }
            in_deliverables = False
            in_breakdown = False
            continue
        
        if not current_task:
            continue
        
        # Estimated Time
        if re.match(r'^- \*\*Estimated Time\*\*:', line):
            time_match = re.search(r': (.+?)$', line)
            if time_match:
                current_task['estimated_time'] = time_match.group(1).strip()
            in_deliverables = False
            in_breakdown = False
            continue
        
        # Priority
        if re.match(r'^- \*\*Priority\*\*:', line):
            priority_match = re.search(r': (.+?)$', line)
            if priority_match:
                current_task['priority'] = priority_match.group(1).strip()
            in_deliverables = False
            in_breakdown = False
            continue
        
        # Breakdown header
        if re.match(r'^- \*\*Breakdown\*\*:', line):
            in_breakdown = True
            in_deliverables = False
            continue
        
        # Deliverables header
        if re.match(r'^- \*\*Deliverables\*\*:', line):
            in_deliverables = True
            in_breakdown = False
            continue
        
        # Breakdown items (indented with 2 spaces)
        if in_breakdown and re.match(r'^  -', line):
            breakdown_match = re.match(r'^  - (.+?)$', line)
            if breakdown_match:
                current_task['breakdown'].append(breakdown_match.group(1).strip())
            continue
        
        # Deliverables items (indented with 2 spaces, comes after deliverables header)
        if in_deliverables and re.match(r'^  -', line):
            deliverable_match = re.match(r'^  - (.+?)$', line)
            if deliverable_match:
                current_task['deliverables'].append(deliverable_match.group(1).strip())
            continue
    
    # Add last task and phase
    if current_task:
        current_phase['tasks'].append(current_task)
    if current_phase:
        phases.append(current_phase)
    
    return phases

def parse_days_from_string(time_str):
    """Parse time string like '2-4 days' or '3 days' - already in days"""
    if not time_str:
        return 'N/A'
    
    if 'days' in time_str.lower():
        # Already in days, just return as is
        return time_str
    
    # Extract numbers from time string (for backward compatibility)
    numbers = re.findall(r'\d+(?:\.\d+)?', time_str)
    if not numbers:
        return time_str
    
    if len(numbers) == 1:
        return f"{numbers[0]} days"
    elif len(numbers) >= 2:
        return f"{numbers[0]}-{numbers[1]} days"
    else:
        return time_str

def parse_days_to_numbers(time_str):
    """Parse time string like '2-4 days' to tuple of (min, max) integers"""
    if not time_str:
        return (0, 0)
    
    numbers = re.findall(r'\d+', time_str)
    if not numbers:
        return (0, 0)
    
    if len(numbers) == 1:
        days = int(numbers[0])
        return (days, days)
    elif len(numbers) >= 2:
        return (int(numbers[0]), int(numbers[1]))
    else:
        return (0, 0)

def create_excel_file(phases, output_file):
    """Create an organized Excel file with multiple sheets"""
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Summary
        summary_data = {
            'Phase': [],
            'Phase Name': [],
            'Min Days': [],
            'Max Days': [],
            'Avg Days': [],
            'Tasks Count': []
        }
        
        total_min = 0
        total_max = 0
        
        for phase in phases:
            summary_data['Phase'].append(f"Phase {phase['phase_num']}")
            summary_data['Phase Name'].append(phase['phase_name'])
            # Values are already in days
            min_days = phase['min_hours']  # Actually days
            max_days = phase['max_hours']   # Actually days
            summary_data['Min Days'].append(min_days)
            summary_data['Max Days'].append(max_days)
            avg_days = (min_days + max_days) / 2
            summary_data['Avg Days'].append(round(avg_days, 1))
            summary_data['Tasks Count'].append(len(phase['tasks']))
            total_min += min_days
            total_max += max_days
        
        # Add totals row
        summary_data['Phase'].append('TOTAL')
        summary_data['Phase Name'].append('')
        summary_data['Min Days'].append(total_min)
        summary_data['Max Days'].append(total_max)
        summary_data['Avg Days'].append(round((total_min + total_max) / 2, 1))
        summary_data['Tasks Count'].append(sum(len(p['tasks']) for p in phases))
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: All Tasks
        tasks_data = {
            'Phase': [],
            'Task ID': [],
            'Task Name': [],
            'Estimated Time': [],
            'Priority': [],
            'Breakdown Items': [],
            'Deliverables': []
        }
        
        for phase in phases:
            for task in phase['tasks']:
                tasks_data['Phase'].append(f"Phase {phase['phase_num']}")
                tasks_data['Task ID'].append(task['task_id'])
                tasks_data['Task Name'].append(task['task_name'])
                estimated_days = parse_days_from_string(task['estimated_time'])
                tasks_data['Estimated Time'].append(estimated_days)
                tasks_data['Priority'].append(task['priority'])
                breakdown_text = '\n• '.join(task['breakdown']) if task['breakdown'] else 'N/A'
                if breakdown_text != 'N/A':
                    breakdown_text = '• ' + breakdown_text
                tasks_data['Breakdown Items'].append(breakdown_text)
                deliverables_text = '\n• '.join(task['deliverables']) if task['deliverables'] else 'N/A'
                if deliverables_text != 'N/A':
                    deliverables_text = '• ' + deliverables_text
                tasks_data['Deliverables'].append(deliverables_text)
        
        df_tasks = pd.DataFrame(tasks_data)
        df_tasks.to_excel(writer, sheet_name='All Tasks', index=False)
        
        # Individual sheets for each phase
        for phase in phases:
            phase_tasks_data = {
                'Task ID': [],
                'Task Name': [],
                'Estimated Time': [],
                'Priority': [],
                'Breakdown': [],
                'Deliverables': []
            }
            
            for task in phase['tasks']:
                phase_tasks_data['Task ID'].append(task['task_id'])
                phase_tasks_data['Task Name'].append(task['task_name'])
                estimated_days = parse_days_from_string(task['estimated_time'])
                phase_tasks_data['Estimated Time'].append(estimated_days)
                phase_tasks_data['Priority'].append(task['priority'])
                breakdown_text = '\n• '.join(task['breakdown']) if task['breakdown'] else 'N/A'
                if breakdown_text != 'N/A':
                    breakdown_text = '• ' + breakdown_text
                phase_tasks_data['Breakdown'].append(breakdown_text)
                deliverables_text = '\n• '.join(task['deliverables']) if task['deliverables'] else 'N/A'
                if deliverables_text != 'N/A':
                    deliverables_text = '• ' + deliverables_text
                phase_tasks_data['Deliverables'].append(deliverables_text)
            
            df_phase = pd.DataFrame(phase_tasks_data)
            sheet_name = f"Phase {phase['phase_num']}"
            # Excel sheet names have a 31 character limit
            if len(sheet_name) > 31:
                sheet_name = f"P{phase['phase_num']}"
            df_phase.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Format the Excel file
    format_excel_file(output_file, phases)

def format_excel_file(file_path, phases):
    """Apply formatting to the Excel file"""
    
    wb = load_workbook(file_path)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format total row
        total_row = ws.max_row
        for cell in ws[total_row]:
            cell.fill = total_fill
            cell.font = total_font
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        # Apply borders and alignment
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
                if cell.column in [3, 4, 5, 6]:  # Number columns (Days)
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                    # Format as number with 1 decimal place
                    if cell.row > 1 and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'
    
    # Format All Tasks sheet
    if 'All Tasks' in wb.sheetnames:
        ws = wb['All Tasks']
        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 60
        
        # Set row heights for wrapped text
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 60
        
        # Apply borders and wrap text
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
                if cell.column == 4:  # Priority column
                    if 'High' in str(cell.value):
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    elif 'Medium' in str(cell.value):
                        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                    elif 'Low' in str(cell.value):
                        cell.fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
    
    # Format individual phase sheets
    for phase in phases:
        sheet_name = f"Phase {phase['phase_num']}"
        if len(sheet_name) > 31:
            sheet_name = f"P{phase['phase_num']}"
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Format header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 60
            
            # Set row heights
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 80
            
            # Apply borders and wrap text
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.border = border
                    if cell.row > 1:
                        cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
                    if cell.column == 4:  # Priority column
                        if 'High' in str(cell.value):
                            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        elif 'Medium' in str(cell.value):
                            cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                        elif 'Low' in str(cell.value):
                            cell.fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
    
    wb.save(file_path)

if __name__ == '__main__':
    input_file = 'estimates.md'
    output_file = 'project_estimates.xlsx'
    
    print(f"Parsing {input_file}...")
    phases = parse_estimates_md(input_file)
    print(f"Found {len(phases)} phases with {sum(len(p['tasks']) for p in phases)} total tasks")
    
    print(f"Creating Excel file: {output_file}...")
    create_excel_file(phases, output_file)
    print(f"Excel file created successfully: {output_file}")
    print(f"\nFile location: {output_file}")

